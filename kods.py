import os, sys, time

from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

from openpyxl import Workbook, load_workbook
wb = Workbook()
ws = wb.active

if os.path.exists("datne.xlsx") == False:
    print ("\nDatne nav atrasta, lūdzu ievadiet savus datus (šo procesu vairs nebūs jāatkārto).")

    ws["A2"] = "ORTUS"
    ws["B1"] = "Lietotājvārds"
    ws["C1"] = "Parole"
    ws["D1"] = "Novirziens"
    ws["E1"] = "Kurss"
    ws["F1"] = "Grupa"

    ws["B2"] = input("ORTUS lietotājvārds: ")
    ws["C2"] = input("Parole: ")

    print("\nNākamā sadaļa paredzēta Jūsu grafika automātiskai atvēršanai.")
    ws["D2"] = input("Programmas kods (piemēram, RDBD0): ")
    ws["E2"] = input("Kurss: ")
    ws["F2"] = input("Grupa: ")

    wb.save(filename="datne.xlsx")

wb = load_workbook("datne.xlsx")
ws = wb.active

ortus = "https://id2.rtu.lv/openam/UI/Login?module=LDAP&locale=lv"
estudijas = "https://id2.rtu.lv/openam/UI/Login?locale=lv&goto=https%3A%2F%2Festudijas.rtu.lv"
nodarbibas = "https://nodarbibas.rtu.lv"

ortus_name = ws["B2"].value
ortus_pass = ws["C2"].value
novirziens = ws["D2"].value
kurss = ws["E2"].value
grupa = ws["F2"].value

while (True):
    print("\nKurai saitei vēlaties pieslēgties? Lūdzu ievadiet attiecīgo ciparu!")
    inputs = input("ORTUS - 1 \nE-studijas - 2 \nNodarbību grafiks - 3 \nDzēst vai pārrakstīt datus - 4 \nApturēt programmu - 5 \n")

    if inputs == "1":
        service = Service()
        option = webdriver.ChromeOptions()
        option.add_experimental_option('excludeSwitches', ['enable-logging'])
        option.add_experimental_option("detach", True)
        option.add_argument("--log-level=3")
        driver = webdriver.Chrome(service=service, options=option)
        driver.get(ortus)
        driver.maximize_window()
        time.sleep(2)

        username = driver.find_element(By.ID, "IDToken1")
        username.send_keys(str(ortus_name))
        password = driver.find_element(By.ID, "IDToken2")
        password.send_keys(str(ortus_pass))
        button = driver.find_element(By.NAME, "Login.Submit")
        button.click()

    elif inputs == "2":
        service = Service()
        option = webdriver.ChromeOptions()
        option.add_experimental_option('excludeSwitches', ['enable-logging'])
        option.add_experimental_option("detach", True)
        option.add_argument("--log-level=3")
        driver = webdriver.Chrome(service=service, options=option)
        driver.get(estudijas)
        driver.maximize_window()
        time.sleep(2)

        username = driver.find_element(By.ID, "IDToken1")
        username.send_keys(str(ortus_name))
        password = driver.find_element(By.ID, "IDToken2")
        password.send_keys(str(ortus_pass))
        button = driver.find_element(By.NAME, "Login.Submit")
        button.click()

    elif inputs == "3":
        service = Service()
        option = webdriver.ChromeOptions()
        option.add_experimental_option('excludeSwitches', ['enable-logging'])
        option.add_experimental_option("detach", True)
        option.add_argument("--log-level=3")
        driver = webdriver.Chrome(service=service, options=option)
        driver.get(nodarbibas)
        driver.maximize_window()
        time.sleep(2)

        button1 = driver.find_element(By.CLASS_NAME, "filter-option-inner-inner")
        button1.click()
        element = driver.switch_to.active_element
        element.send_keys(str(novirziens))
        element.send_keys(Keys.ENTER)

        button2 = driver.find_element(By.ID, "course-id")
        button2.click()
        element = driver.switch_to.active_element
        element.send_keys(str(kurss))
        element.send_keys(Keys.ENTER)
        time.sleep(0.1)

        button3 = driver.find_element(By.ID, "group-id")
        button3.click()
        element = driver.switch_to.active_element
        element.send_keys(str(grupa))
        element.send_keys(Keys.ENTER)

    elif inputs == "4":
        os.remove("datne.xlsx")
        print("\nDatne tika dzēsta, lūdzu palaidiet programmu vēlreiz, lai atjaunotu savus datus (variet droši aiztaisīt logu)!")
        time.sleep(2)
        sys.exit()

    elif inputs == "5":
        print("\nUz redzēšanos (variet droši aiztaisīt logu)!")
        time.sleep(2)
        sys.exit()
    else:
        print("\nKļūda, lūdzu mēģiniet vēlreiz.")